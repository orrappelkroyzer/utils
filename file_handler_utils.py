import sys
from pathlib import Path
local_python_path = str(Path(__file__).parents[1])
if local_python_path not in sys.path:
   sys.path.append(local_python_path)
from utils.utils import load_config, get_logger
logger = get_logger(__name__)
config = load_config(Path(local_python_path) / "config.json")

import json

import pandas as pd
from openpyxl import load_workbook

from utils.lock import file_lock


OVERWRITE_FILE = "overwrite_file"
OVERWRITE_SHEET = "overwrite_sheet"
APPEND_SHEET = "append_sheet"


def write_csv(df, filename, output_dir=None, index=False):
    if output_dir is None:
        output_dir = config["output_dir"]
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    fn = output_dir / "{}.csv".format(filename)
    logger.info("Writing csv to {}".format(fn))
    lock_path = Path(f"{fn}.lock")
    with file_lock(lock_path=lock_path, timeout_seconds=600.0):
        df.to_csv(fn, index=index)


def read_csv(file_path, **kwargs):
    file_path = Path(file_path)
    logger.info(f"Reading csv from {file_path}")
    lock_path = Path(f"{file_path}.lock")
    with file_lock(lock_path=lock_path, timeout_seconds=600.0):
        return pd.read_csv(file_path, **kwargs)


def write_json(data, filename, output_dir=None, ensure_ascii=False, indent=2):
    if output_dir is None:
        output_dir = config["output_dir"]
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    fn = output_dir / "{}.json".format(filename)
    logger.info("Writing json to {}".format(fn))
    lock_path = Path(f"{fn}.lock")
    with file_lock(lock_path=lock_path, timeout_seconds=600.0):
        with fn.open("w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=ensure_ascii, indent=indent)


def read_json(file_path, **kwargs):
    file_path = Path(file_path)
    logger.info(f"Reading json from {file_path}")
    lock_path = Path(f"{file_path}.lock")
    encoding = kwargs.pop("encoding", "utf-8")
    with file_lock(lock_path=lock_path, timeout_seconds=600.0):
        with file_path.open("r", encoding=encoding) as f:
            return json.load(f, **kwargs)


def write_excel(
    df,
    filename,
    output_dir=None,
    sheet_name="Sheet1",
    index=False,
    override_mode=OVERWRITE_FILE,
):
    if output_dir is None:
        output_dir = config["output_dir"]
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    fn = output_dir / "{}.xlsx".format(filename)
    logger.info(f"Writing excel to sheet {sheet_name} in file {fn}")
    lock_path = Path(f"{fn}.lock")
    with file_lock(lock_path=lock_path, timeout_seconds=600.0):
        if not fn.exists() or override_mode == OVERWRITE_FILE:
            df.to_excel(fn, sheet_name=sheet_name, index=index)
            return
        if override_mode == OVERWRITE_SHEET:
            try:
                with pd.ExcelWriter(
                    fn, engine="openpyxl", mode="a", if_sheet_exists="replace"
                ) as writer:  # type: ignore
                    df.to_excel(writer, sheet_name=sheet_name, index=index)
                return
            except TypeError:
                try:
                    wb = load_workbook(fn)
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        wb.remove(ws)
                        wb.save(fn)
                except Exception as e:
                    logger.warning(
                        f"Failed to remove existing sheet '{sheet_name}' from {fn}: {e}"
                    )
                with pd.ExcelWriter(fn, engine="openpyxl", mode="a") as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=index)
                return
        if override_mode == APPEND_SHEET:
            with pd.ExcelWriter(fn, engine="openpyxl", mode="a") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=index)
                return
        raise AssertionError(f"received illegal override_mode {override_mode}")


def read_excel(file_path, **kwargs):
    file_path = Path(file_path)
    logger.info(f"Reading excel from {file_path}")
    lock_path = Path(f"{file_path}.lock")
    with file_lock(lock_path=lock_path, timeout_seconds=600.0):
        return pd.read_excel(file_path, **kwargs)


def append_row_to_excel(df_row: pd.DataFrame, excel_path: Path, sheet_name: str = "Sheet1"):
    if not excel_path.exists():
        df_row.to_excel(excel_path, sheet_name=sheet_name, index=False)
        return

    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        headers = list(df_row.columns)
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
    else:
        ws = wb[sheet_name]

    next_row = ws.max_row + 1
    for col_idx, col_name in enumerate(df_row.columns, start=1):
        value = df_row[col_name].iloc[0]
        if hasattr(value, "tolist"):
            value = value.tolist()
        ws.cell(row=next_row, column=col_idx, value=value)

    wb.save(excel_path)
