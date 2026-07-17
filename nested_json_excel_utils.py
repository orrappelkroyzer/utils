import sys
import json
import re
from collections.abc import Callable
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

local_python_path = str(Path(__file__).parents[1])
if local_python_path not in sys.path:
    sys.path.append(local_python_path)

from utils.file_handler_utils import read_excel, read_json, write_excel
from utils.utils import get_logger, load_config

logger = get_logger(__name__)
config = load_config(Path(local_python_path) / "config.json")

MARKDOWN_LINK_PATTERN = re.compile(r"\[([^\]]+)\]\([^)]+\)")
EXCEL_DATA_START_ROW = 3
LONG_TEXT_COLUMN_THRESHOLD = 30
MIN_LONG_TEXT_COLUMN_WIDTH = 24
MAX_EXCEL_COLUMN_WIDTH = 60
EXCEL_COLUMN_WIDTH_PADDING = 2


def ensure_record_list(extracted_json) -> list[dict]:
    if isinstance(extracted_json, list):
        return extracted_json
    if isinstance(extracted_json, dict):
        return [extracted_json]
    raise RuntimeError(f"Expected extracted output to be list/dict, got {type(extracted_json)}")


def clean_markdown_link_text(value: str) -> str:
    cleaned_value = value.strip()
    full_match = MARKDOWN_LINK_PATTERN.fullmatch(cleaned_value)
    if full_match:
        return full_match.group(1)
    return MARKDOWN_LINK_PATTERN.sub(r"\1", cleaned_value)


def normalize_value_for_excel(value):
    if isinstance(value, str):
        return clean_markdown_link_text(value)
    if isinstance(value, dict):
        return {
            key: normalize_value_for_excel(nested_value)
            for key, nested_value in value.items()
        }
    if isinstance(value, list):
        normalized_items = [normalize_value_for_excel(item) for item in value]
        return deduplicate_list_values(normalized_items)
    return value


def get_list_item_dedup_key(item) -> str | int | float | bool | None:
    if isinstance(item, dict):
        return json.dumps(item, ensure_ascii=False, sort_keys=True)
    if isinstance(item, list):
        return json.dumps(item, ensure_ascii=False, sort_keys=True)
    if isinstance(item, str):
        return item.strip()
    return item


def deduplicate_list_values(values: list) -> list:
    deduplicated_values: list = []
    seen_keys: set = set()
    for value in values:
        dedup_key = get_list_item_dedup_key(value)
        if dedup_key in seen_keys:
            continue
        seen_keys.add(dedup_key)
        deduplicated_values.append(value)
    return deduplicated_values


def prepare_record_for_excel(
    record: dict,
    record_normalizer: Callable[[dict], dict] | None = None,
) -> dict:
    prepared_record = dict(record)
    if record_normalizer is not None:
        prepared_record = record_normalizer(prepared_record)
    return normalize_value_for_excel(prepared_record)


def load_json_records_from_directory(results_dir: Path) -> list[dict]:
    if not results_dir.is_dir():
        raise NotADirectoryError(f"JSON results directory not found: {results_dir}")

    result_paths = sorted(results_dir.glob("*.json"))
    if not result_paths:
        raise FileNotFoundError(f"No JSON result files found in {results_dir}")

    records: list[dict] = []
    for result_path in result_paths:
        result_data = read_json(result_path)
        result_records = ensure_record_list(result_data)
        records.extend(result_records)
        logger.info(f"Loaded {len(result_records)} record(s) from {result_path.name}")
    return records


def is_list_of_dicts(value) -> bool:
    return isinstance(value, list) and bool(value) and all(isinstance(item, dict) for item in value)


def get_row_expanded_fields_for_record(record: dict, schema_fields: set[str]) -> list[str]:
    expanded_fields: list[str] = []
    for field_name, value in record.items():
        if is_list_of_dicts(value):
            expanded_fields.append(field_name)
            continue
        if isinstance(value, list) and not value and field_name in schema_fields:
            expanded_fields.append(field_name)
    return expanded_fields


def collect_row_expanded_field_prefixes(
    records: list[dict],
    schema_fields: set[str],
    record_normalizer: Callable[[dict], dict] | None = None,
) -> set[str]:
    expanded_field_prefixes: set[str] = set()
    for record in records:
        normalized_record = prepare_record_for_excel(record, record_normalizer=record_normalizer)
        expanded_field_prefixes.update(
            get_row_expanded_fields_for_record(normalized_record, schema_fields)
        )
    return expanded_field_prefixes


def get_subfields_for_list_field(items: list) -> list[str]:
    subfields: list[str] = []
    seen_subfields: set[str] = set()
    for item in items:
        if not isinstance(item, dict):
            continue
        for subfield_name in item:
            if subfield_name in seen_subfields:
                continue
            seen_subfields.add(subfield_name)
            subfields.append(subfield_name)
    return subfields


def is_single_word_text(value: str) -> bool:
    stripped_value = value.strip()
    if not stripped_value:
        return False
    return not re.search(r"\s", stripped_value)


def format_list_for_excel(values: list) -> str:
    formatted_values: list[str] = []
    for value in values:
        if isinstance(value, (dict, list)):
            formatted_values.append(json.dumps(value, ensure_ascii=False))
        else:
            formatted_values.append(str(value))

    if formatted_values and all(is_single_word_text(text) for text in formatted_values):
        return ",".join(formatted_values)
    return ",\n".join(formatted_values)


def format_list_of_dicts_for_excel(values: list[dict]) -> str:
    return "\n".join(json.dumps(value, ensure_ascii=False) for value in values)


def flatten_dict_item_for_excel(item: dict, parent_key: str) -> dict:
    flattened_item: dict = {}
    for key, value in item.items():
        column_name = f"{parent_key}.{key}"
        if isinstance(value, list):
            flattened_item[column_name] = format_list_for_excel(value)
        elif isinstance(value, dict):
            flattened_item[column_name] = json.dumps(value, ensure_ascii=False)
        else:
            flattened_item[column_name] = value
    return flattened_item


def empty_expanded_field_columns(field_name: str, subfields: list[str]) -> dict:
    if not subfields:
        return {field_name: None}
    return {f"{field_name}.{subfield}": None for subfield in subfields}


def flatten_result_record(
    record: dict,
    parent_key: str = "",
    row_expanded_fields: set[str] | None = None,
) -> dict:
    row_expanded_fields = row_expanded_fields or set()
    flattened_record: dict = {}
    for key, value in record.items():
        if not parent_key and key in row_expanded_fields:
            continue
        column_name = f"{parent_key}.{key}" if parent_key else str(key)
        if isinstance(value, dict):
            if value:
                flattened_record.update(
                    flatten_result_record(value, column_name, row_expanded_fields)
                )
            else:
                flattened_record[column_name] = None
        elif isinstance(value, list):
            if value and all(isinstance(item, dict) for item in value):
                flattened_record[column_name] = format_list_of_dicts_for_excel(value)
            else:
                flattened_record[column_name] = format_list_for_excel(value)
        else:
            flattened_record[column_name] = value
    return flattened_record


def split_record_into_excel_rows(
    record: dict,
    schema_fields: set[str],
    record_normalizer: Callable[[dict], dict] | None = None,
) -> list[dict]:
    normalized_record = prepare_record_for_excel(record, record_normalizer=record_normalizer)
    expanded_fields = get_row_expanded_fields_for_record(normalized_record, schema_fields)
    base_record = {
        key: value
        for key, value in normalized_record.items()
        if key not in expanded_fields
    }
    base_flat = flatten_result_record(
        base_record,
        row_expanded_fields=set(expanded_fields),
    )

    if not expanded_fields:
        return [base_flat]

    row_count = max(
        max(len(normalized_record.get(field_name, []) or []) for field_name in expanded_fields),
        1,
    )
    expanded_rows: list[dict] = []
    for row_index in range(row_count):
        row = dict(base_flat)
        for field_name in expanded_fields:
            field_items = normalized_record.get(field_name, [])
            if not isinstance(field_items, list):
                field_items = []
            dict_items = [item for item in field_items if isinstance(item, dict)]
            subfields = get_subfields_for_list_field(dict_items)
            if row_index < len(dict_items):
                row.update(flatten_dict_item_for_excel(dict_items[row_index], field_name))
            else:
                row.update(empty_expanded_field_columns(field_name, subfields))
            for subfield in subfields:
                row.setdefault(f"{field_name}.{subfield}", None)
        expanded_rows.append(row)
    return expanded_rows


def subcolumn_sort_key(column_name: str) -> tuple:
    parts = column_name.split(".")[1:]
    sort_parts: list[tuple] = []
    subfield_order = {"min": 0, "max": 1, "raw_text": 2}
    for part in parts:
        if part.isdigit():
            sort_parts.append((0, int(part)))
        elif part in subfield_order:
            sort_parts.append((1, subfield_order[part], part))
        else:
            sort_parts.append((2, part))
    return tuple(sort_parts)


def order_columns_for_nested_headers(column_names: list[str]) -> list[str]:
    primary_order: list[str] = []
    groups: dict[str, list[str]] = {}
    for column_name in column_names:
        primary_level = column_name.split(".", maxsplit=1)[0]
        if primary_level not in groups:
            primary_order.append(primary_level)
            groups[primary_level] = []
        groups[primary_level].append(column_name)

    ordered_columns: list[str] = []
    for primary_level in primary_order:
        group_columns = groups[primary_level]
        if len(group_columns) == 1:
            ordered_columns.append(group_columns[0])
        else:
            ordered_columns.extend(sorted(group_columns, key=subcolumn_sort_key))
    return ordered_columns


def records_to_excel_dataframe(
    records: list[dict],
    row_expanded_schema_fields: set[str],
    record_normalizer: Callable[[dict], dict] | None = None,
) -> tuple[pd.DataFrame, list[tuple[int, int]], set[str]]:
    excel_rows: list[dict] = []
    row_groups: list[tuple[int, int]] = []
    current_row = 0

    for record in records:
        record_rows = split_record_into_excel_rows(
            record=record,
            schema_fields=row_expanded_schema_fields,
            record_normalizer=record_normalizer,
        )
        if len(record_rows) > 1:
            row_groups.append((current_row, current_row + len(record_rows) - 1))
        excel_rows.extend(record_rows)
        current_row += len(record_rows)

    records_df = pd.DataFrame(excel_rows)
    ordered_columns = order_columns_for_nested_headers(list(records_df.columns))
    row_expanded_field_prefixes = collect_row_expanded_field_prefixes(
        records=records,
        schema_fields=row_expanded_schema_fields,
        record_normalizer=record_normalizer,
    )
    return records_df.reindex(columns=ordered_columns), row_groups, row_expanded_field_prefixes


def get_merge_column_indices(
    column_names: list[str],
    row_expanded_field_prefixes: set[str],
) -> list[int]:
    merge_column_indices: list[int] = []
    for column_index, column_name in enumerate(column_names, start=1):
        primary_level = column_name.split(".", maxsplit=1)[0]
        if primary_level not in row_expanded_field_prefixes:
            merge_column_indices.append(column_index)
    return merge_column_indices


def merge_excel_row_groups(
    excel_path: Path,
    row_groups: list[tuple[int, int]],
    column_names: list[str],
    row_expanded_field_prefixes: set[str],
    data_start_row: int = EXCEL_DATA_START_ROW,
) -> None:
    if not row_groups:
        return

    workbook = load_workbook(excel_path)
    worksheet = workbook.active
    merge_column_indices = get_merge_column_indices(
        column_names=column_names,
        row_expanded_field_prefixes=row_expanded_field_prefixes,
    )

    for start_row, end_row in row_groups:
        if start_row >= end_row:
            continue
        excel_start_row = start_row + data_start_row
        excel_end_row = end_row + data_start_row
        for column_index in merge_column_indices:
            worksheet.merge_cells(
                start_row=excel_start_row,
                start_column=column_index,
                end_row=excel_end_row,
                end_column=column_index,
            )

    workbook.save(excel_path)
    workbook.close()


def has_nested_excel_headers(excel_path: Path) -> bool:
    workbook = load_workbook(excel_path, read_only=False)
    worksheet = workbook.active
    merged_ranges = list(worksheet.merged_cells.ranges)
    workbook.close()
    return any(
        merged_range.min_row == 1
        and (
            merged_range.max_row == 2
            or (merged_range.max_row == 1 and merged_range.max_col > merged_range.min_col)
        )
        for merged_range in merged_ranges
    )


def read_results_excel(excel_path: Path) -> pd.DataFrame:
    if not has_nested_excel_headers(excel_path):
        return read_excel(excel_path)

    results_df = read_excel(excel_path, header=[0, 1])
    results_df.columns = [
        primary_level
        if str(secondary_level).startswith("Unnamed")
        else f"{primary_level}.{secondary_level}"
        for primary_level, secondary_level in results_df.columns
    ]
    return results_df


def parse_column_header_levels(column_name: str) -> tuple[str, str | None]:
    primary_level, *secondary_levels = column_name.split(".", maxsplit=1)
    secondary_level = secondary_levels[0] if secondary_levels else None
    return primary_level, secondary_level


def group_column_header_levels(column_names: list[str]) -> list[tuple[str, int, int, bool]]:
    parsed_columns = [parse_column_header_levels(column_name) for column_name in column_names]
    groups: list[tuple[str, int, int, bool]] = []
    column_index = 0
    while column_index < len(parsed_columns):
        primary_level = parsed_columns[column_index][0]
        start_column = column_index + 1
        next_column_index = column_index + 1
        while (
            next_column_index < len(parsed_columns)
            and parsed_columns[next_column_index][0] == primary_level
        ):
            next_column_index += 1
        has_secondary_level = any(
            parsed_columns[group_column_index][1] is not None
            for group_column_index in range(column_index, next_column_index)
        )
        groups.append((primary_level, start_column, next_column_index, has_secondary_level))
        column_index = next_column_index
    return groups


def format_nested_excel_headers(excel_path: Path, column_names: list[str]) -> None:
    workbook = load_workbook(excel_path)
    worksheet = workbook.active
    worksheet.insert_rows(1)

    parsed_columns = [parse_column_header_levels(column_name) for column_name in column_names]
    for column_index, (primary_level, secondary_level) in enumerate(parsed_columns, start=1):
        worksheet.cell(row=1, column=column_index, value=primary_level)
        worksheet.cell(row=2, column=column_index, value=secondary_level)

    for primary_level, start_column, end_column, has_secondary_level in group_column_header_levels(
        column_names
    ):
        if end_column > start_column:
            worksheet.merge_cells(
                start_row=1,
                start_column=start_column,
                end_row=1,
                end_column=end_column,
            )
        elif not has_secondary_level:
            worksheet.merge_cells(
                start_row=1,
                start_column=start_column,
                end_row=2,
                end_column=start_column,
            )

    workbook.save(excel_path)
    workbook.close()


def get_cell_display_length(cell_value) -> int:
    if cell_value is None:
        return 0
    text = str(cell_value).strip()
    if not text:
        return 0
    return max(len(line) for line in text.splitlines())


def get_column_width_for_length(max_length: int) -> float | None:
    if max_length < LONG_TEXT_COLUMN_THRESHOLD:
        return None
    target_width = max_length + EXCEL_COLUMN_WIDTH_PADDING
    return float(min(max(target_width, MIN_LONG_TEXT_COLUMN_WIDTH), MAX_EXCEL_COLUMN_WIDTH))


def apply_excel_cell_formatting(excel_path: Path, data_start_row: int = EXCEL_DATA_START_ROW) -> None:
    workbook = load_workbook(excel_path)
    worksheet = workbook.active
    max_column = worksheet.max_column
    column_lengths = [0] * max_column

    for row in worksheet.iter_rows(min_row=data_start_row, max_col=max_column):
        for cell in row:
            if cell.column is None:
                continue
            column_index = cell.column - 1
            if isinstance(cell.value, str) and "\n" in cell.value:
                cell.alignment = cell.alignment.copy(wrap_text=True)
            display_length = get_cell_display_length(cell.value)
            if display_length > column_lengths[column_index]:
                column_lengths[column_index] = display_length

    for column_index, max_length in enumerate(column_lengths, start=1):
        column_width = get_column_width_for_length(max_length)
        if column_width is not None:
            worksheet.column_dimensions[get_column_letter(column_index)].width = column_width

    workbook.save(excel_path)
    workbook.close()


def write_records_to_excel(
    records: list[dict],
    filename: str,
    output_dir: Path,
    row_expanded_schema_fields: set[str] | None = None,
    record_normalizer: Callable[[dict], dict] | None = None,
) -> Path:
    schema_fields = row_expanded_schema_fields or set()
    records_df, row_groups, row_expanded_field_prefixes = records_to_excel_dataframe(
        records=records,
        row_expanded_schema_fields=schema_fields,
        record_normalizer=record_normalizer,
    )
    excel_path = output_dir / f"{filename}.xlsx"
    write_excel(
        records_df,
        filename=filename,
        output_dir=output_dir,
    )
    format_nested_excel_headers(
        excel_path=excel_path,
        column_names=list(records_df.columns),
    )
    merge_excel_row_groups(
        excel_path=excel_path,
        row_groups=row_groups,
        column_names=list(records_df.columns),
        row_expanded_field_prefixes=row_expanded_field_prefixes,
    )
    apply_excel_cell_formatting(excel_path=excel_path)
    return excel_path
